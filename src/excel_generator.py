"""
Módulo Generador de Excel
Crea archivo Excel con formato profesional
"""

import pandas as pd
import os
from datetime import datetime
import config
from src.logger import logger

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl no disponible - formato limitado")


class ExcelGenerator:
    """Genera archivo Excel con formato"""
    
    def __init__(self):
        """Inicializa el generador"""
        self.archivo_salida = None
    
    def generar_nombre_archivo(self):
        """
        Genera nombre para el archivo de salida
        
        Returns:
            String con nombre de archivo
        """
        timestamp = datetime.now().strftime(config.FORMATO_ARCHIVO)
        nombre = f"{config.PREFIJO_OUTPUT}_{timestamp}.xlsx"
        return os.path.join(config.DIR_OUTPUT, nombre)
    
    def crear_hoja_resumen(self, writer, stats):
        """
        Crea hoja de resumen con estadísticas
        
        Args:
            writer: ExcelWriter
            stats: Dict con estadísticas
        """
        if not config.GENERAR_HOJA_RESUMEN:
            return
        
        resumen_data = {
            'Métrica': [
                'Fecha de Procesamiento',
                'Total Empleados',
                'Total Registros',
                'Turnos Completos',
                'Turnos Incompletos',
                'Duplicados Eliminados',
                'Estados Inferidos',
                'Errores',
                'Advertencias'
            ],
            'Valor': [
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                stats.get('empleados_unicos', 0),
                stats.get('total_registros', 0),
                stats.get('turnos_completos', 0),
                stats.get('turnos_incompletos', 0),
                stats.get('duplicados_eliminados', 0),
                stats.get('estados_inferidos', 0),
                stats.get('errores', 0),
                stats.get('advertencias', 0)
            ]
        }
        
        df_resumen = pd.DataFrame(resumen_data)
        df_resumen.to_excel(writer, sheet_name='Resumen', index=False)
        
        logger.info("Hoja de resumen creada")
    
    def aplicar_formato(self, ruta_archivo, nombre_hoja='Reporte'):
        """
        Aplica formato al archivo Excel
        
        Args:
            ruta_archivo: Ruta al archivo
            nombre_hoja: Nombre de la hoja a formatear
        """
        if not OPENPYXL_AVAILABLE:
            logger.warning("No se puede aplicar formato - openpyxl no disponible")
            return
        
        try:
            wb = load_workbook(ruta_archivo)
            
            # Formatear la hoja dada
            if nombre_hoja in wb.sheetnames:
                ws = wb[nombre_hoja]
                
                # Colores
                color_encabezado = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                color_ok = PatternFill(start_color=config.COLORES['VERDE'][1:], end_color=config.COLORES['VERDE'][1:], fill_type='solid')
                color_advertencia = PatternFill(start_color=config.COLORES['AMARILLO'][1:], end_color=config.COLORES['AMARILLO'][1:], fill_type='solid')
                color_error = PatternFill(start_color=config.COLORES['NARANJA'][1:], end_color=config.COLORES['NARANJA'][1:], fill_type='solid')
                color_nocturno = PatternFill(start_color=config.COLORES['AZUL'][1:], end_color=config.COLORES['AZUL'][1:], fill_type='solid')
                color_morado = PatternFill(start_color=config.COLORES['MORADO'][1:], end_color=config.COLORES['MORADO'][1:], fill_type='solid')
                
                # Fuentes
                font_encabezado = Font(bold=True, color='FFFFFF', size=11)
                font_normal = Font(size=10)
                
                # Alineación
                align_center = Alignment(horizontal='center', vertical='center')
                align_left = Alignment(horizontal='left', vertical='center')
                
                # Bordes
                border_delgado = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Nombres de columnas de esta hoja
                headers = []
                for cell in ws[1]:
                    headers.append(cell.value)
                
                # Aplicar formato a encabezados
                for col_num, column_title in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.fill = color_encabezado
                    cell.font = font_encabezado
                    cell.alignment = align_center
                    cell.border = border_delgado
                    
                    # Ajustar ancho de columna
                    col_letter = get_column_letter(col_num)
                    ancho = config.ANCHOS_COLUMNAS.get(column_title, 20) if nombre_hoja == 'Reporte' else 25
                    if column_title == 'OBSERVACION': ancho = 50
                    if column_title == 'NOMBRE COMPLETO DEL COLABORADOR': ancho = 35
                    ws.column_dimensions[col_letter].width = ancho
                
                # Buscar indice de columna observacion
                obs_col_idx = None
                for i, header in enumerate(headers, 1):
                    if header == 'OBSERVACION':
                        obs_col_idx = i
                        break
                
                # Aplicar formato a datos
                for row_num in range(2, ws.max_row + 1):
                    # Obtener observación si existe
                    observacion = ''
                    if obs_col_idx:
                        obs_cell = ws.cell(row=row_num, column=obs_col_idx)
                        observacion = obs_cell.value or ''
                    
                    # Determinar color de fila
                    fill_color = None
                    if 'OK' in observacion or observacion == config.OBSERVACIONES.get('OK', 'Sin observaciones'):
                        fill_color = color_ok
                    elif observacion == config.OBSERVACIONES.get('SIN_REGISTROS', 'SIN REGISTROS'):
                        fill_color = color_nocturno
                    elif config.OBSERVACIONES.get('SALIDA_ESTANDAR_NOCTURNA') in observacion:
                        fill_color = color_morado
                    elif 'TURNO_NOCTURNO' in observacion:
                        fill_color = color_nocturno
                    elif 'ALERTA' in observacion:
                        fill_color = color_error
                    elif observacion and observacion != config.OBSERVACIONES.get('OK', 'Sin observaciones'):
                        fill_color = color_advertencia
                    
                    # Aplicar formato a todas las celdas de la fila
                    for col_num in range(1, len(headers) + 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.font = font_normal
                        cell.border = border_delgado
                        
                        # Alineación según columna (centrar numéricas)
                        if isinstance(cell.value, (int, float)) or headers[col_num-1] in ['SEMANA DEL AÑO', 'CANTIDAD EMPLEADOS']:
                            cell.alignment = align_center
                        else:
                            cell.alignment = align_left
                        
                        # Color de fondo
                        if fill_color:
                            cell.fill = fill_color
                
                # Congelar primera fila
                ws.freeze_panes = 'A2'
                
                # Agregar Data Validation si es la hoja Reporte y existe la columna OBSERVACIONES_1
                if nombre_hoja == 'Reporte' and 'OBSERVACIONES_1' in headers and 'Conceptos' in wb.sheetnames:
                    obs1_col_idx = headers.index('OBSERVACIONES_1') + 1
                    obs1_letter = get_column_letter(obs1_col_idx)
                    ws_conceptos = wb['Conceptos']
                    max_row_conceptos = ws_conceptos.max_row
                    if max_row_conceptos > 1:
                        dv = DataValidation(type="list", formula1=f"Conceptos!$A$2:$A${max_row_conceptos}", allow_blank=True)
                        # Optional styling for the dropdown
                        dv.error ='Su valor no está en la lista'
                        dv.errorTitle = 'Entrada inválida'
                        dv.prompt = 'Seleccione una observación'
                        dv.promptTitle = 'Observación'
                        ws.add_data_validation(dv)
                        dv.add(f'{obs1_letter}2:{obs1_letter}{max(2, ws.max_row)}')
                    
                    # Ocultar la hoja de conceptos para que quede limpio el reporte
                    wb['Conceptos'].sheet_state = 'hidden'
            
            # Guardar
            wb.save(ruta_archivo)
            logger.info(f"Formato aplicado exitosamente a {nombre_hoja}")
            
        except Exception as e:
            logger.error(f"Error al aplicar formato en {nombre_hoja}: {str(e)}")
            
    def crear_hoja_empleados(self, writer, df_resultado):
        """
        Crea hoja agrupada por empleado y horas por semana
        """
        if df_resultado.empty:
            return
            
        try:
            df = df_resultado.copy()
            # Calcular fecha datetime y extraer semana
            df['FECHA_DT'] = pd.to_datetime(df['FECHA'], format=config.FORMATO_FECHA_OUTPUT)
            df['SEMANA'] = df['FECHA_DT'].dt.isocalendar().week
            
            # Convertir horas a numerico
            df['HORAS_NUM'] = pd.to_numeric(df['TOTAL HORAS LABORADAS'], errors='coerce').fillna(0)
            
            group_cols = ['CODIGO COLABORADOR', 'NOMBRE COMPLETO DEL COLABORADOR', 'CARGO', 'SEMANA']
            if 'LIMITE_HORAS_SEMANA' in df.columns:
                group_cols.append('LIMITE_HORAS_SEMANA')
                
            # Agrupar
            df_agrupado = df.groupby(group_cols).agg(
                TOTAL_HORAS_SEMANA=('HORAS_NUM', 'sum')
            ).reset_index()
            
            # Redondear
            df_agrupado['TOTAL_HORAS_SEMANA'] = df_agrupado['TOTAL_HORAS_SEMANA'].round(2)
            
            # Generar observacion
            if 'LIMITE_HORAS_SEMANA' in df_agrupado.columns:
                def calcular_obs(row):
                    if row['TOTAL_HORAS_SEMANA'] > row['LIMITE_HORAS_SEMANA']:
                        return f"ALERTA: EXCEDE LÍMITE SEMANAL ({row['LIMITE_HORAS_SEMANA']} hrs)"
                    return config.OBSERVACIONES['OK']
                df_agrupado['OBSERVACION'] = df_agrupado.apply(calcular_obs, axis=1)
                
            # Renombrar
            renames = {'SEMANA': 'SEMANA DEL AÑO', 'LIMITE_HORAS_SEMANA': 'LÍMITE HORAS SEMANA', 'TOTAL_HORAS_SEMANA': 'TOTAL HORAS SEMANA'}
            df_agrupado = df_agrupado.rename(columns=renames)
            
            # Guardar en excel
            df_agrupado.to_excel(writer, sheet_name='Horas por Empleado', index=False)
            logger.info("Hoja de 'Horas por Empleado' creada")
        except Exception as e:
            logger.error(f"Error al generar hoja Horas por Empleado: {str(e)}")

    def crear_hoja_cargos(self, writer, df_resultado):
        """
        Crea hoja agrupada por cargo y cuenta alertas
        """
        if df_resultado.empty or 'CARGO' not in df_resultado.columns:
            return
            
        try:
            df = df_resultado.copy()
            # Contar exceso de limites
            def tiene_alerta_exceso(obs):
                return 1 if isinstance(obs, str) and 'EXCEDE LÍMITE DE HORAS DEL CARGO' in obs else 0
                
            df['ALERTA_EXCESO'] = df['OBSERVACION'].apply(tiene_alerta_exceso)
            
            group_cols = ['CARGO']
            if 'COLABORADORES_ESPERADOS' in df.columns:
                group_cols.append('COLABORADORES_ESPERADOS')
                
            # Agrupar por cargo
            df_agrupado = df.groupby(group_cols).agg(
                CANTIDAD_EMPLEADOS=('CODIGO COLABORADOR', 'nunique'),
                DIAS_CON_ALERTA_EXCESO=('ALERTA_EXCESO', 'sum')
            ).reset_index()
            
            # Generar observacion
            if 'COLABORADORES_ESPERADOS' in df_agrupado.columns:
                def calcular_obs(row):
                    if pd.notna(row['COLABORADORES_ESPERADOS']) and row['COLABORADORES_ESPERADOS'] != '':
                        try:
                            esperados = float(row['COLABORADORES_ESPERADOS'])
                            actuales = float(row['CANTIDAD_EMPLEADOS'])
                            if actuales > esperados:
                                return f"ALERTA: SOBRECUPO ({actuales} vs {esperados} esperados)"
                            elif actuales < esperados:
                                return f"ALERTA: FALTAN EMPLEADOS ({actuales} vs {esperados} esperados)"
                        except ValueError:
                            pass
                    return config.OBSERVACIONES['OK']
                df_agrupado['OBSERVACION'] = df_agrupado.apply(calcular_obs, axis=1)
                
            renames = {'CANTIDAD_EMPLEADOS': 'COLABORADORES REALES', 'COLABORADORES_ESPERADOS': 'COLABORADORES ESPERADOS'}
            df_agrupado = df_agrupado.rename(columns=renames)
            
            # Filtrar cargos vacíos
            df_agrupado = df_agrupado[df_agrupado['CARGO'] != '']
            
            # Guardar en excel
            df_agrupado.to_excel(writer, sheet_name='Resumen por Cargo', index=False)
            logger.info("Hoja de 'Resumen por Cargo' creada")
        except Exception as e:
            logger.error(f"Error al generar hoja Resumen por Cargo: {str(e)}")
    
    def generar_excel(self, df_resultado, stats=None, df_conceptos=None):
        """
        Genera archivo Excel con los resultados

        Args:
            df_resultado: DataFrame con resultados finales
            stats: Dict con estadísticas (opcional)
            df_conceptos: DataFrame con columna 'observaciones' para el dropdown
                          de OBSERVACIONES_1. Si es None se omite la hoja.

        Returns:
            Ruta al archivo generado
        """
        logger.log_fase("GENERACIÓN DE ARCHIVO EXCEL")
        
        # Crear directorio de salida si no existe
        os.makedirs(config.DIR_OUTPUT, exist_ok=True)
        
        # Generar nombre de archivo
        ruta_salida = self.generar_nombre_archivo()
        
        # Crear archivo Excel
        with pd.ExcelWriter(ruta_salida, engine='openpyxl') as writer:
            # Hoja principal
            df_resultado.to_excel(writer, sheet_name='Reporte', index=False)
            
            # Nuevas Hojas de Agrupación
            self.crear_hoja_empleados(writer, df_resultado)
            self.crear_hoja_cargos(writer, df_resultado)
            
            # Hoja de Conceptos (para validación de datos en OBSERVACIONES_1)
            if df_conceptos is not None and 'observaciones' in df_conceptos.columns:
                try:
                    df_conceptos[['observaciones']].dropna().to_excel(writer, sheet_name='Conceptos', index=False)
                except Exception as e:
                    logger.warning(f"No se pudo crear hoja de Conceptos: {str(e)}")
            
            # Hoja de resumen
            if stats:
                self.crear_hoja_resumen(writer, stats)
        
        # Aplicar formato
        self.aplicar_formato(ruta_salida, 'Reporte')
        self.aplicar_formato(ruta_salida, 'Horas por Empleado')
        self.aplicar_formato(ruta_salida, 'Resumen por Cargo')
        
        self.archivo_salida = ruta_salida
        
        logger.info(config.MENSAJES['excel_generado'])
        logger.info(f"Archivo guardado en: {ruta_salida}")
        
        return ruta_salida
    
    def generar_casos_especiales(self, df_resultado):
        """
        Genera archivo con casos que requieren revisión manual
        
        Args:
            df_resultado: DataFrame con resultados
            
        Returns:
            Ruta al archivo o None
        """
        if not config.GENERAR_CASOS_ESPECIALES:
            return None
        
        # Filtrar casos especiales
        casos = df_resultado[
            (df_resultado['OBSERVACION'].str.contains('ALERTA', na=False)) |
            (df_resultado['OBSERVACION'].str.contains('REQUIERE', na=False)) |
            (df_resultado['OBSERVACION'].str.contains('INDEFINIDO', na=False)) |
            (df_resultado['OBSERVACION'].str.contains('DATOS_CORRUPTOS', na=False))
        ].copy()
        
        if len(casos) == 0:
            logger.info("No hay casos especiales para revisar")
            return None
        
        # Generar archivo
        timestamp = datetime.now().strftime(config.FORMATO_ARCHIVO)
        nombre = f"CASOS_REVISION_{timestamp}.xlsx"
        ruta = os.path.join(config.DIR_OUTPUT, nombre)
        
        casos.to_excel(ruta, index=False)
        
        logger.info(f"Archivo de casos especiales generado: {ruta}")
        logger.info(f"Total casos para revisión: {len(casos)}")
        
        return ruta
