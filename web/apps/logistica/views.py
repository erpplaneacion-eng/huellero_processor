"""
Views para el área de Logística
"""

import os
from datetime import datetime
from pathlib import Path

from django.conf import settings
from django.core.management import call_command
from django.http import HttpResponse, JsonResponse, FileResponse, Http404
from django.views import View
from django.views.generic import TemplateView
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator

from .models import RegistroAsistencia
from .processor import HuelleroProcessor


@method_decorator(login_required, name='dispatch')
class IndexView(TemplateView):
    """Vista principal del área de Logística"""
    template_name = 'logistica/index.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['area'] = settings.AREAS_CONFIG.get('logistica', {})
        context['area_key'] = 'logistica'
        return context


@method_decorator([login_required, csrf_exempt], name='dispatch')
class ProcesarView(View):
    """API para procesar archivo de huellero"""

    def post(self, request):
        try:
            # Validar que se envió un archivo
            if 'archivo' not in request.FILES:
                return JsonResponse({
                    'success': False,
                    'error': 'No se envió ningún archivo'
                }, status=400)

            archivo = request.FILES['archivo']

            if archivo.name == '':
                return JsonResponse({
                    'success': False,
                    'error': 'No se seleccionó ningún archivo'
                }, status=400)

            # Validar extensión
            if not archivo.name.lower().endswith(('.xls', '.xlsx')):
                return JsonResponse({
                    'success': False,
                    'error': 'Formato de archivo no válido. Use .xls o .xlsx'
                }, status=400)

            # Guardar archivo temporalmente
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            extension = os.path.splitext(archivo.name)[1]
            nombre_archivo = f"huellero_logistica_{timestamp}{extension}"
            ruta_archivo = settings.DATA_INPUT_DIR / nombre_archivo

            # Guardar el archivo
            with open(ruta_archivo, 'wb+') as destino:
                for chunk in archivo.chunks():
                    destino.write(chunk)

            # Obtener opción de maestro
            usar_maestro = request.POST.get('usar_maestro', 'true').lower() == 'true'

            # Procesar archivo
            processor = HuelleroProcessor(area='logistica')
            resultado = processor.procesar(str(ruta_archivo), usar_maestro)

            return JsonResponse(resultado)

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error durante el procesamiento: {str(e)}'
            }, status=500)


@method_decorator([login_required, csrf_exempt], name='dispatch')
class GuardarObs1View(View):
    """
    Guarda el valor de OBSERVACIONES_1 seleccionado por el usuario en el dashboard.
    POST /logistica/api/registros/obs1/
    Body JSON: {"registro_id": 5, "obs1": "texto seleccionado"}
    """

    def post(self, request):
        import json
        try:
            data = json.loads(request.body)
            registro_id = int(data.get('registro_id', 0))
            obs1 = str(data.get('obs1', '')).strip()

            registro = RegistroAsistencia.objects.get(pk=registro_id)
            registro.observaciones_1 = obs1
            registro.save(update_fields=['observaciones_1', 'actualizado_en'])

            return JsonResponse({'ok': True})

        except RegistroAsistencia.DoesNotExist:
            return JsonResponse({'ok': False, 'error': 'Registro no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'ok': False, 'error': str(e)}, status=500)


@method_decorator(login_required, name='dispatch')
class ListarRegistrosView(View):
    """
    Devuelve registros guardados en BD paginados por empleado (10 por página).
    GET /logistica/api/registros/?page=1
    """
    PAGE_SIZE = 10

    def get(self, request):
        try:
            from apps.logistica.models import Concepto
            from django.db.models import Q

            page     = max(1, int(request.GET.get('page', 1)))
            search   = request.GET.get('search', '').strip()
            mes      = request.GET.get('mes', '').strip()   # "YYYY-MM"
            filtrado = bool(search or mes)

            processor = HuelleroProcessor(area='logistica')
            codigos_excluidos = processor._cargar_codigos_excluidos()

            base_qs = RegistroAsistencia.objects.exclude(codigo__in=codigos_excluidos)

            # Totales globales (para el header)
            total_registros_global = base_qs.count()
            total_empleados_global = base_qs.values('codigo').distinct().count()

            if filtrado:
                # ── Filtro servidor: devuelve todos los coincidentes sin paginar ──
                filtered_qs = base_qs

                if mes:
                    try:
                        year, month = mes.split('-')
                        filtered_qs = filtered_qs.filter(
                            fecha__year=int(year), fecha__month=int(month)
                        )
                    except (ValueError, AttributeError):
                        pass

                if search:
                    q_filter = Q(nombre__icontains=search) | Q(documento__icontains=search)
                    try:
                        q_filter |= Q(codigo=int(search))
                    except ValueError:
                        pass
                    filtered_qs = filtered_qs.filter(q_filter)

                codigos_pagina = list(
                    filtered_qs.values_list('codigo', flat=True).distinct().order_by('codigo')
                )
                registros_qs = filtered_qs.filter(
                    codigo__in=codigos_pagina
                ).order_by('codigo', 'fecha', 'hora_ingreso')
                has_more = False
                total_empleados = len(codigos_pagina)

            else:
                # ── Sin filtro: paginación normal ──
                codigos_todos = list(
                    base_qs.values_list('codigo', flat=True).distinct().order_by('codigo')
                )
                total_empleados = len(codigos_todos)
                offset = (page - 1) * self.PAGE_SIZE
                codigos_pagina = codigos_todos[offset:offset + self.PAGE_SIZE]
                has_more = (offset + self.PAGE_SIZE) < total_empleados
                registros_qs = base_qs.filter(
                    codigo__in=codigos_pagina
                ).order_by('codigo', 'fecha', 'hora_ingreso')

            horarios_por_codigo = processor._cargar_horarios_por_codigo()

            def _best_fit_turno(codigo_int, hora_ingreso_str):
                turnos_raw = horarios_por_codigo.get(codigo_int, [])
                if not turnos_raw or not hora_ingreso_str or ':' not in hora_ingreso_str:
                    return ''
                try:
                    parts = hora_ingreso_str.split(':')
                    ingreso_min = int(parts[0]) * 60 + int(parts[1])
                    e, s = min(turnos_raw, key=lambda t: abs(t[0] - ingreso_min))
                    return f"{e//60:02d}:{e%60:02d}-{s//60:02d}:{s%60:02d}"
                except Exception:
                    return ''

            empleados = {}
            for r in registros_qs:
                codigo = str(r.codigo)
                if codigo not in empleados:
                    empleados[codigo] = {
                        'codigo':    codigo,
                        'nombre':    r.nombre,
                        'documento': r.documento,
                        'cargo':     r.cargo,
                        'registros': [],
                    }
                empleados[codigo]['registros'].append({
                    'id':          r.id,
                    'fecha':       r.fecha.strftime('%d/%m/%Y'),
                    'dia':         r.dia,
                    'am':          r.marcaciones_am,
                    'pm':          r.marcaciones_pm,
                    'ingreso':     r.hora_ingreso,
                    'salida':      r.hora_salida,
                    'horas':       r.total_horas,
                    'limite':      r.limite_horas_dia,
                    'observacion': r.observacion,
                    'obs1':        r.observaciones_1,
                    'turno':       _best_fit_turno(r.codigo, r.hora_ingreso),
                })

            datos = list(empleados.values())
            conceptos = list(
                Concepto.objects.values_list('observaciones', flat=True).order_by('observaciones')
            ) if page == 1 else []

            return JsonResponse({
                'success':       True,
                'datos':         datos,
                'conceptos':     conceptos,
                'archivo':       None,
                'archivo_casos': None,
                'desde_db':      True,
                'page':          page,
                'has_more':      has_more,
                'stats': {
                    'empleados_unicos':      total_empleados_global,
                    'total_registros':       total_registros_global,
                    'turnos_completos':      0,
                    'turnos_incompletos':    0,
                    'duplicados_eliminados': 0,
                    'estados_inferidos':     0,
                },
            })

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)


@method_decorator(login_required, name='dispatch')
class DescargarRegistrosExcelView(View):
    """
    Genera y descarga un Excel con los registros de BD filtrados por rango de fechas.
    GET /logistica/api/registros/excel/?desde=YYYY-MM-DD&hasta=YYYY-MM-DD
    """

    # Mismos colores que el dashboard y el ExcelGenerator del pipeline
    _FILL = {
        'verde':   'C6EFCE',
        'amarillo':'FFEB9C',
        'naranja': 'FFC7CE',
        'azul':    'DDEBF7',
        'morado':  'E1BEE7',
        'gris':    'D9D9D9',
    }

    @staticmethod
    def _color_observacion(obs):
        obs = obs or ''
        if not obs or obs in ('Sin observaciones', 'OK'):
            return 'verde'
        if 'SIN REGISTROS' in obs or 'SIN_REGISTROS' in obs:
            return 'gris'
        if 'SALIDA_ESTANDAR_NOCTURNA' in obs or 'Salida Inferida Estándar' in obs:
            return 'morado'
        if 'TURNO_NOCTURNO' in obs or 'Turno nocturno' in obs:
            return 'azul'
        if 'ALERTA' in obs.upper():
            return 'naranja'
        return 'amarillo'

    def get(self, request):
        import io
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        try:
            desde_str = request.GET.get('desde', '')
            hasta_str = request.GET.get('hasta', '')

            processor = HuelleroProcessor(area='logistica')
            codigos_excluidos = processor._cargar_codigos_excluidos()

            qs = RegistroAsistencia.objects.exclude(
                codigo__in=codigos_excluidos
            ).order_by('codigo', 'fecha', 'hora_ingreso')

            if desde_str:
                qs = qs.filter(fecha__gte=desde_str)
            if hasta_str:
                qs = qs.filter(fecha__lte=hasta_str)

            registros = list(qs)

            # ── Estilos comunes ───────────────────────────────────────────
            fill_header = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            font_header = Font(bold=True, color='FFFFFF', size=11)
            font_normal = Font(size=10)
            align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
            align_left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
            border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'),  bottom=Side(style='thin'),
            )
            fills = {k: PatternFill(start_color=v, end_color=v, fill_type='solid')
                     for k, v in self._FILL.items()}

            wb = Workbook()

            # ══════════════════════════════════════════════════════════════
            # Hoja 1: Registros
            # ══════════════════════════════════════════════════════════════
            ws = wb.active
            ws.title = 'Registros'
            ws.freeze_panes = 'A2'

            HEADERS = [
                'CODIGO', 'NOMBRE COMPLETO', 'DOCUMENTO', 'CARGO',
                'FECHA', 'DÍA', 'HORA INGRESO', 'HORA SALIDA',
                'MARC. AM', 'MARC. PM', 'TOTAL HORAS', 'LÍMITE HORAS',
                'OBSERVACION', 'OBSERVACIONES_1',
            ]
            ANCHOS = [10, 35, 14, 25, 12, 12, 13, 13, 10, 10, 12, 12, 50, 30]

            for col, (h, ancho) in enumerate(zip(HEADERS, ANCHOS), 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill, cell.font, cell.alignment, cell.border = fill_header, font_header, align_center, border
                ws.column_dimensions[get_column_letter(col)].width = ancho

            for row_num, r in enumerate(registros, 2):
                valores = [
                    r.codigo, r.nombre, r.documento, r.cargo,
                    r.fecha.strftime('%d/%m/%Y'), r.dia,
                    r.hora_ingreso, r.hora_salida,
                    r.marcaciones_am, r.marcaciones_pm,
                    r.total_horas, r.limite_horas_dia,
                    r.observacion, r.observaciones_1,
                ]
                fill = fills[self._color_observacion(r.observacion)]
                for col, valor in enumerate(valores, 1):
                    cell = ws.cell(row=row_num, column=col, value=valor)
                    cell.fill, cell.font, cell.border = fill, font_normal, border
                    cell.alignment = align_center if isinstance(valor, (int, float)) else align_left

            # ══════════════════════════════════════════════════════════════
            # Hoja 2: Horas por Semana
            # ══════════════════════════════════════════════════════════════
            self._crear_hoja_horas_semana(
                wb, registros,
                fill_header, font_header, font_normal,
                align_center, align_left, border,
            )

            buffer = io.BytesIO()
            wb.save(buffer)
            data = buffer.getvalue()

            desde_label = desde_str.replace('-', '') if desde_str else 'inicio'
            hasta_label  = hasta_str.replace('-', '')  if hasta_str  else 'fin'
            filename = f'registros_{desde_label}_{hasta_label}.xlsx'

            response = HttpResponse(
                data,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        except Exception as e:
            import traceback
            return HttpResponse(
                f'Error: {e}\n\n{traceback.format_exc()}',
                status=500, content_type='text/plain',
            )

    def _crear_hoja_horas_semana(self, wb, registros, fill_header, font_header,
                                  font_normal, align_center, align_left, border):
        """Hoja 2: horas acumuladas por empleado por semana vs límite del cargo."""
        from datetime import timedelta
        from openpyxl.styles import PatternFill
        from apps.logistica.models import Cargo, Empleado

        # ── Mapa cargo_id → horas_semana ─────────────────────────────────
        limite_por_cargo_id = {
            c.id_cargo: c.horas_semana
            for c in Cargo.objects.all()
        }
        # Mapa codigo_empleado → cargo_id (para empleados que tienen cargo asignado)
        cargo_id_por_codigo = {
            e.codigo: e.cargo_id
            for e in Empleado.objects.exclude(cargo__isnull=True).select_related('cargo')
        }

        # ── Acumular horas por (codigo, semana_iso) ───────────────────────
        from collections import defaultdict
        semanas = defaultdict(lambda: {
            'nombre': '', 'cargo': '', 'horas': 0.0,
            'fecha_inicio': None, 'anio': 0, 'semana': 0,
        })

        for r in registros:
            if not r.fecha or r.total_horas is None:
                continue
            iso = r.fecha.isocalendar()
            key = (r.codigo, iso[0], iso[1])   # (codigo, año_iso, semana_iso)
            d = semanas[key]
            d['nombre']  = d['nombre']  or r.nombre
            d['cargo']   = d['cargo']   or r.cargo
            d['horas']  += float(r.total_horas or 0)
            d['anio']    = iso[0]
            d['semana']  = iso[1]
            # Lunes de esa semana
            if d['fecha_inicio'] is None:
                lunes = r.fecha - timedelta(days=r.fecha.weekday())
                d['fecha_inicio'] = lunes

        # ── Construir filas ordenadas ─────────────────────────────────────
        filas = []
        for (codigo, anio, semana), d in sorted(semanas.items()):
            cargo_id  = cargo_id_por_codigo.get(codigo)
            limite    = limite_por_cargo_id.get(cargo_id) if cargo_id else None
            horas     = round(d['horas'], 2)
            diferencia = round(horas - limite, 2) if limite else None
            fecha_ini = d['fecha_inicio']
            fecha_fin = (fecha_ini + timedelta(days=6)) if fecha_ini else None

            if limite is None:
                estado, color = 'SIN LÍMITE', 'D9D9D9'
            elif horas > limite:
                estado, color = 'EXCEDE', 'FFC7CE'
            else:
                estado, color = 'OK', 'C6EFCE'

            filas.append((
                codigo,
                d['nombre'],
                d['cargo'] or '',
                limite or '',
                anio,
                semana,
                fecha_ini.strftime('%d/%m/%Y') if fecha_ini else '',
                fecha_fin.strftime('%d/%m/%Y') if fecha_fin else '',
                horas,
                diferencia if diferencia is not None else '',
                estado,
                color,
            ))

        # ── Escribir hoja ─────────────────────────────────────────────────
        ws2 = wb.create_sheet(title='Horas por Semana')
        ws2.freeze_panes = 'A2'

        HEADERS2 = [
            'CODIGO', 'NOMBRE COMPLETO', 'CARGO',
            'LÍMITE HORAS SEMANA', 'AÑO', 'SEMANA',
            'INICIO SEMANA', 'FIN SEMANA',
            'TOTAL HORAS', 'DIFERENCIA', 'ESTADO',
        ]
        ANCHOS2 = [10, 35, 25, 18, 8, 10, 14, 14, 12, 12, 14]

        for col, (h, ancho) in enumerate(zip(HEADERS2, ANCHOS2), 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.fill, cell.font, cell.alignment, cell.border = fill_header, font_header, align_center, border
            from openpyxl.utils import get_column_letter
            ws2.column_dimensions[get_column_letter(col)].width = ancho

        for row_num, fila in enumerate(filas, 2):
            *valores, color = fila
            fill_row = PatternFill(start_color=color, end_color=color, fill_type='solid')
            for col, valor in enumerate(valores, 1):
                cell = ws2.cell(row=row_num, column=col, value=valor)
                cell.fill, cell.font, cell.border = fill_row, font_normal, border
                cell.alignment = align_center if isinstance(valor, (int, float)) else align_left


@method_decorator(login_required, name='dispatch')
class DescargarView(View):
    """Vista para descargar archivos generados"""

    def get(self, request, filename):
        ruta_archivo = settings.DATA_OUTPUT_DIR / filename

        if not ruta_archivo.exists():
            raise Http404('Archivo no encontrado')

        return FileResponse(
            open(ruta_archivo, 'rb'),
            as_attachment=True,
            filename=filename
        )


@csrf_exempt
def cron_sincronizar_planta(request):
    """
    GET/POST /logistica/cron/sincronizar-planta/?token=<WEBHOOK_SECRET_TOKEN>
    Ejecuta sincronizar_planta: copia registros nuevos de tabla_planta → maestro_empleado.
    Pensado para ser llamado diariamente por Railway Cron.
    """
    token_esperado = os.environ.get('WEBHOOK_SECRET_TOKEN', '')
    token_recibido = (
        request.GET.get('token') or
        request.headers.get('Authorization', '').removeprefix('Bearer ').strip()
    )
    if not token_esperado or token_recibido != token_esperado:
        return JsonResponse({'error': 'No autorizado'}, status=401)

    import io
    from django.core.management import call_command

    buffer = io.StringIO()
    try:
        call_command('sincronizar_planta', stdout=buffer)
        return JsonResponse({'success': True, 'log': buffer.getvalue()})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
